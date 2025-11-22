#!/usr/bin/env python3
"""
build_vuln_register_from_yaml.py

Builds a Vulnerability Register (CSV/XLSX) from a YAML definition file.

This is intended for SMEs as part of a Vulnerability & Patch Management toolkit.
It lets you maintain a human-readable YAML file and then automatically generate
a structured register that can be used for governance, reporting, and audits.

-------------------------------------------------------------------------------
YAML INPUT FORMAT (EXAMPLE)
-------------------------------------------------------------------------------

vulnerabilities:
  - id: "VULN-001"
    title: "Missing critical security patch on web server"
    description: "Web server WS-01 is missing the latest cumulative security update."
    asset_id: "WS-01"
    asset_name: "Public Web Server 1"
    asset_owner: "Infrastructure Team"
    environment: "Production"
    business_impact: "Customer-facing website outage / data exposure risk."
    discovery_date: "2025-01-15"
    source: "Internal scan (Nessus)"
    cvss_score: 8.1
    severity: "High"
    status: "Open"
    remediation_owner: "John Doe"
    remediation_target_date: "2025-01-25"
    remediation_actual_date: ""
    remediation_plan: "Apply January cumulative patch and reboot during maintenance window."
    exception_id: ""
    risk_acceptance: "No"
    verification_status: "Not Verified"
    notes: "Prioritise ahead of monthly patch cycle."
    related_tickets: "CHG-12345;INC-5555"
    standard_mapping: "ISO27001 A.8.8; CIS 7.1; NIST CSF PR.IP-12"

  - id: "VULN-002"
    title: "Outdated TLS configuration on legacy app"
    description: "Legacy app supports TLS 1.0 and weak cipher suites."
    asset_id: "APP-LEG-01"
    asset_name: "Legacy App Server"
    asset_owner: "App Team"
    environment: "UAT"
    business_impact: "Potential downgrade / confidentiality risk."
    discovery_date: "2025-01-10"
    source: "Manual review"
    cvss_score: 5.6
    severity: "Medium"
    status: "In Progress"
    remediation_owner: "Jane Smith"
    remediation_target_date: "2025-02-01"
    remediation_actual_date: ""
    remediation_plan: "Disable TLS 1.0/1.1 and weak ciphers; retest."
    exception_id: "EXC-2025-001"
    risk_acceptance: "Temporary until vendor upgrade"
    verification_status: "Pending"
    notes: ""
    related_tickets: "CHG-12348"
    standard_mapping: "ISO27001 A.8.24; NIST CSF PR.DS-2"


Fields are flexible: any extra fields will be ignored unless mapped explicitly.
The following fields are recognised and mapped into the register:

    - id
    - title
    - description
    - asset_id
    - asset_name
    - asset_owner
    - environment
    - business_impact
    - discovery_date
    - source
    - cvss_score
    - severity
    - status
    - remediation_owner
    - remediation_target_date
    - remediation_actual_date
    - remediation_plan
    - exception_id
    - risk_acceptance
    - verification_status
    - notes
    - related_tickets
    - standard_mapping

-------------------------------------------------------------------------------
USAGE
-------------------------------------------------------------------------------

    python build_vuln_register_from_yaml.py --input vuln_register.yaml

Optional arguments:

    --csv-output   Path to CSV output (default: vulnerability_register.csv)
    --xlsx-output  Path to XLSX output (requires openpyxl; default: none)

Examples:

    # CSV only
    python build_vuln_register_from_yaml.py -i vuln_register.yaml

    # CSV + XLSX
    python build_vuln_register_from_yaml.py -i vuln_register.yaml \
        --csv-output data/vulnerability_register.csv \
        --xlsx-output data/vulnerability_register.xlsx

Requirements:

    pip install pyyaml openpyxl
"""

import argparse
import csv
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import yaml
except ImportError:
    yaml = None  # type: ignore

try:
    from openpyxl import Workbook  # type: ignore
except ImportError:
    Workbook = None  # type: ignore


@dataclass
class VulnerabilityRecord:
    id: str
    title: str
    description: str
    asset_id: str
    asset_name: str
    asset_owner: str
    environment: str
    business_impact: str
    discovery_date: str
    source: str
    cvss_score: str
    severity: str
    status: str
    remediation_owner: str
    remediation_target_date: str
    remediation_actual_date: str
    remediation_plan: str
    exception_id: str
    risk_acceptance: str
    verification_status: str
    notes: str
    related_tickets: str
    standard_mapping: str


FIELDNAMES = [
    "id",
    "title",
    "description",
    "asset_id",
    "asset_name",
    "asset_owner",
    "environment",
    "business_impact",
    "discovery_date",
    "source",
    "cvss_score",
    "severity",
    "status",
    "remediation_owner",
    "remediation_target_date",
    "remediation_actual_date",
    "remediation_plan",
    "exception_id",
    "risk_acceptance",
    "verification_status",
    "notes",
    "related_tickets",
    "standard_mapping",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a vulnerability register (CSV/XLSX) from a YAML definition file."
    )
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        type=Path,
        help="Path to YAML file containing 'vulnerabilities' list.",
    )
    parser.add_argument(
        "--csv-output",
        "-c",
        type=Path,
        default=Path("vulnerability_register.csv"),
        help="Path to CSV output (default: vulnerability_register.csv).",
    )
    parser.add_argument(
        "--xlsx-output",
        "-x",
        type=Path,
        help="Optional path to XLSX output (requires openpyxl).",
    )
    return parser.parse_args()


def ensure_yaml_available() -> None:
    global yaml
    if yaml is None:
        raise SystemExit(
            "PyYAML is required but not installed. Install it with:\n\n"
            "    pip install pyyaml\n"
        )


def ensure_openpyxl_available() -> None:
    global Workbook
    if Workbook is None:
        raise SystemExit(
            "openpyxl is required for XLSX output but is not installed. Install it with:\n\n"
            "    pip install openpyxl\n"
        )


def load_yaml_vulnerabilities(path: Path) -> List[Dict[str, Any]]:
    ensure_yaml_available()

    if not path.exists():
        raise FileNotFoundError(f"YAML file not found: {path}")

    with path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    if not isinstance(data, dict) or "vulnerabilities" not in data:
        raise ValueError(
            f"YAML file must contain a top-level 'vulnerabilities' key pointing to a list."
        )

    vulns = data["vulnerabilities"]

    if not isinstance(vulns, list):
        raise ValueError("'vulnerabilities' must be a list of items.")

    return vulns


def normalise_vulnerability(item: Dict[str, Any]) -> VulnerabilityRecord:
    def get_str(key: str) -> str:
        value = item.get(key, "")
        if value is None:
            return ""
        return str(value)

    return VulnerabilityRecord(
        id=get_str("id"),
        title=get_str("title"),
        description=get_str("description"),
        asset_id=get_str("asset_id"),
        asset_name=get_str("asset_name"),
        asset_owner=get_str("asset_owner"),
        environment=get_str("environment"),
        business_impact=get_str("business_impact"),
        discovery_date=get_str("discovery_date"),
        source=get_str("source"),
        cvss_score=get_str("cvss_score"),
        severity=get_str("severity"),
        status=get_str("status"),
        remediation_owner=get_str("remediation_owner"),
        remediation_target_date=get_str("remediation_target_date"),
        remediation_actual_date=get_str("remediation_actual_date"),
        remediation_plan=get_str("remediation_plan"),
        exception_id=get_str("exception_id"),
        risk_acceptance=get_str("risk_acceptance"),
        verification_status=get_str("verification_status"),
        notes=get_str("notes"),
        related_tickets=get_str("related_tickets"),
        standard_mapping=get_str("standard_mapping"),
    )


def write_csv(records: List[VulnerabilityRecord], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for r in records:
            writer.writerow(asdict(r))


def write_xlsx(records: List[VulnerabilityRecord], path: Path) -> None:
    ensure_openpyxl_available()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerability Register"

    # Header row
    ws.append(FIELDNAMES)

    # Data rows
    for r in records:
        row = [getattr(r, field) for field in FIELDNAMES]
        ws.append(row)

    # Auto-fit-ish: set column width based on max length (simple heuristic)
    for col_idx, field in enumerate(FIELDNAMES, start=1):
        max_len = len(field)
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    wb.save(path)


def main() -> int:
    args = parse_args()

    try:
        raw_vulns = load_yaml_vulnerabilities(args.input)
    except Exception as e:
        print(f"Error loading YAML vulnerabilities: {e}")
        return 1

    records = [normalise_vulnerability(item) for item in raw_vulns]

    print(f"Loaded {len(records)} vulnerability record(s) from {args.input}")

    # Simple console summary - counts by severity
    severity_counts: Dict[str, int] = {}
    for r in records:
        sev = (r.severity or "").strip() or "Unspecified"
        severity_counts[sev] = severity_counts.get(sev, 0) + 1

    if severity_counts:
        print("Counts by severity:")
        for sev, count in sorted(severity_counts.items(), key=lambda x: x[0]):
            print(f"  {sev}: {count}")

    # Write CSV
    try:
        write_csv(records, args.csv_output)
        print(f"CSV written to: {args.csv_output}")
    except Exception as e:
        print(f"Error writing CSV: {e}")
        return 1

    # Optionally write XLSX
    if args.xlsx_output:
        try:
            write_xlsx(records, args.xlsx_output)
            print(f"XLSX written to: {args.xlsx_output}")
        except Exception as e:
            print(f"Error writing XLSX: {e}")
            return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
